import openpyxl
import pandas as pd
from openpyxl import load_workbook


class WorkbookLoader:
    def __init__(self) -> None:
        self.default_style = { "fontSize": "10pt", "fontFamily": "Tahoma", }

    def get_column_width(self, worksheet):
        columns = [{"index": i} for i in range(0, worksheet.max_column + 1)]
        for col_letter, dim in worksheet.column_dimensions.items():
            try:
                col_index = openpyxl.utils.column_index_from_string(col_letter) - 1
                if dim.width:
                    columns[col_index]["width"] = dim.width*7
            except:
                continue
        return columns

    def get_cell_style(self, cell):
        style = {}
        if cell.font:
            if cell.font.bold:
                style["fontWeight"] = "Bold"
            if cell.font.underline:
                style["textDecoration"] = "Underline"
            if cell.font.color:
                try:
                    if str(cell.font.color.rgb).strip()=="00000000":
                        style["color"] = f"#000"
                    else:
                        style["color"] = f"#{cell.font.color.rgb[2:]}"
                except: 
                    style["color"] = f"#000"

        if cell.fill:
            if cell.fill.start_color.rgb:
                try:
                    if str(cell.fill.start_color.rgb).strip()=="00000000":
                        style["backgroundColor"] = f"#ffffff"
                    else:
                        style["backgroundColor"] = f"#{cell.fill.start_color.rgb[2:]}"
                except: 
                    style["backgroundColor"] = f"#ffffff"

        if cell.border:
            if cell.border.bottom.style:
                style["borderBottom"] = "1px solid #000000"
            if cell.border.left.style:
                style["borderLeft"] = "1px solid #000000"
            if cell.border.right.style:
                style["borderRight"] = "1px solid #000000"
            if cell.border.top.style:
                style["borderTop"] = "1px solid #000000"
        return style
    
    def get_cell_wrap(self, cell):
        if cell.alignment.wrap_text:
            return True
        else:
            return False
        
    def get_cell_format(self, cell):
        format = None
        if cell.number_format:
            format = cell.number_format
        return format

    def get_output_dict(self, file, extension):
        if extension in ["xlsx","xls"]:
            wb = load_workbook(file, data_only=True)
            use_styles = True
        elif extension in ["csv"]:
            df = pd.read_csv(file)
            df = pd.concat([df.columns.to_frame().T, df], ignore_index=True)
            use_styles = False
        else:
            raise ValueError("Unsupported file format. Only .xlsx and .csv are supported.")

        output_dict = {
            "Workbook": {
                "definedNames": [],
                "sheets": [],
                "filterCollection": [],
                "sortCollection": [],
            }
        }
        if use_styles:
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                df = pd.DataFrame(ws.values)

                sheet_dict = {
                    "columns": self.get_column_width(ws),
                    "name": sheet_name,
                    "rows": [
                        {
                            "cells": [
                                { 
                                    "isLocked": False,
                                    "format": self.get_cell_format(cell),
                                    "style": self.get_cell_style(cell) if use_styles else self.default_style,
                                    "value": cell_value,
                                    "wrap": self.get_cell_wrap(cell)
                                }
                                for cell_value, cell in zip(row_values, row) if use_styles
                            ],
                            "height": (ws.row_dimensions[row[0].row].height or 5.25)*(3.24),
                        } 
                        for row_values, row in zip(df.values, ws.iter_rows(min_row=1)) if use_styles
                    ],
                    "usedRange": {"colIndex": len(df.columns) - 1, "rowIndex": len(df)},
                }
                output_dict["Workbook"]["sheets"].append(sheet_dict)
        else:
            sheet_dict = {
                    "name": "Sheet1",
                    "rows": [
                        {
                            "cells": [
                                { 
                                    "isLocked": True,
                                    "style": self.default_style,
                                    "value": cell_value,
                                }
                                for cell_value in row_values
                            ],
                        } 
                        for row_values in df.values
                    ],
                    "usedRange": {"colIndex": len(df.columns) - 1, "rowIndex": len(df)},
                }
            output_dict["Workbook"]["sheets"].append(sheet_dict)
        return output_dict
    

class WorkbookDeLoader:
    def __init__(self) -> None:
        pass

    def get_dataframe(self, output_dict, sheet_number):
        sheet_data = output_dict["sheets"][sheet_number]
        rows = sheet_data["rows"]
        
        data = []
        for row in rows:
            row_data = [cell.get("value", None) for cell in row["cells"]]
            data.append(row_data)
        
        df = pd.DataFrame(data, columns=data[0])
        return df